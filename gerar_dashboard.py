# -*- coding: utf-8 -*-
"""
Gerador automatico do Dashboard OQC - Label Code.

O que faz:
  1) Localiza a planilha de origem a partir do ATALHO (.lnk) que esta nesta pasta.
     Assim, mesmo que a planilha mude de lugar, basta o atalho continuar valido.
  2) Le a aba 'Sheet1', extrai e classifica as inspecoes.
  3) Injeta os dados no 'template.html' e regrava o 'dashboard.html'.

Uso:
  python gerar_dashboard.py
  (ou clique duas vezes em "Atualizar Dashboard OQC.bat")
"""
import os, sys, json, glob, subprocess, datetime, shutil
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("ERRO: a biblioteca 'openpyxl' nao esta instalada.")
    print("Instale com:  pip install openpyxl")
    sys.exit(1)

PASTA = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(PASTA, "template.html")
SAIDA = os.path.join(PASTA, "dashboard.html")
SAIDA_WEB = os.path.join(PASTA, "index.html")   # publicado no GitHub Pages

# Publicacao no GitHub (Pages). O token NAO fica no repositorio:
# e lido deste arquivo local, fora do OneDrive.
REPO_SLUG = "RicardooliveiraLC/LabecodeOQC"
TOKEN_FILE = os.path.join(os.path.expanduser("~"), ".oqc", "github_token.txt")

# ---------------------------------------------------------------------------
def resolver_planilha():
    """Resolve o caminho do .xlsx a partir do atalho .lnk na pasta."""
    lnks = glob.glob(os.path.join(PASTA, "*.lnk"))
    for lnk in lnks:
        ps = ('[Console]::OutputEncoding=[System.Text.Encoding]::UTF8;'
              '$ws=New-Object -ComObject WScript.Shell;'
              '$ws.CreateShortcut("%s").TargetPath' % lnk.replace('"', '`"'))
        try:
            raw = subprocess.run(
                ["powershell", "-NoProfile", "-Command", ps],
                capture_output=True).stdout
            alvo = raw.decode("utf-8", "replace").strip()
        except Exception:
            alvo = ""
        if alvo.lower().endswith(".xlsx") and os.path.exists(alvo):
            print("Planilha (via atalho): %s" % alvo)
            return alvo
    # Fallback: qualquer .xlsx com 'OQC' no nome em pastas conhecidas
    candidatos = []
    for raiz in (PASTA, os.path.dirname(PASTA),
                 os.path.expanduser("~")):
        candidatos += glob.glob(os.path.join(raiz, "*OQC*.xlsx"))
        candidatos += glob.glob(os.path.join(raiz, "*INSPE*OQC*.xlsx"))
    candidatos = [c for c in candidatos if os.path.exists(c)]
    if candidatos:
        alvo = max(candidatos, key=os.path.getmtime)
        print("Planilha (fallback): %s" % alvo)
        return alvo
    print("ERRO: nao foi possivel localizar a planilha (atalho .lnk ausente ou invalido).")
    sys.exit(1)

# ---------------------------------------------------------------------------
CRIT = {
 8:'Diâmetro do tubete',
 10:'Dimensional do produto',
 11:'Impressão conforme arte',
 12:'Versão/Revisão do item',
 13:'Diâmetro externo da bobina',
 14:'Testes de liner/impressão',
 15:'Leitura QR / código de barras',
 16:'Matéria-prima / certificações',
 17:'Produto final conforme amostra',
 18:'Tape test (desplacamento de tinta)',
 19:'Etiquetas de identificação',
 20:'Preenchimento do checklist',
}

def norm_status(v):
    if v is None: return 'NÃO INFORMADO'
    s = str(v).strip().upper()
    if s.startswith('APROVADO'): return 'APROVADO'
    if s.startswith('REPROVADO'): return 'REPROVADO'
    if s.startswith('CONDICIONAL'): return 'CONDICIONAL'
    return 'NÃO INFORMADO'

def is_fail(v):
    if v is None: return False
    s = str(v).strip().upper()
    return ('NÃO CONFORME' in s) or ('SOLICITADO CORRE' in s) or (s == 'CONDICIONAL') or ('PARCIAL' in s)

def extrair(xlsx):
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.worksheets[-1]
    rows = list(ws.iter_rows(values_only=True))
    data = [r for r in rows[1:] if r and r[0] is not None]
    out = []
    for r in data:
        dt = r[1]
        if dt is None:
            continue
        if not hasattr(dt, 'strftime'):
            continue
        fails = [CRIT[c] for c in CRIT if c < len(r) and is_fail(r[c])]
        out.append({
            'id': r[0],
            'date': dt.strftime('%Y-%m-%d'),
            'datetime': dt.strftime('%Y-%m-%d %H:%M'),
            'order': str(r[5]).strip() if len(r) > 5 and r[5] else '',
            'inspector': str(r[6]).strip().title() if len(r) > 6 and r[6] else '',
            'code': str(r[7]).strip() if len(r) > 7 and r[7] else '',
            'status': norm_status(r[22] if len(r) > 22 else None),
            'fails': fails,
            'obs': str(r[21]).strip() if len(r) > 21 and r[21] else '',
        })
    out.sort(key=lambda x: x['datetime'])
    return out

# ---------------------------------------------------------------------------
def _git(*args, check=True):
    return subprocess.run(["git", "-C", PASTA] + list(args),
                          capture_output=True, text=True, encoding="utf-8",
                          errors="replace", check=check)

def publicar_github(msg):
    """Faz commit e push para o GitHub (se for repo e houver token)."""
    if not os.path.isdir(os.path.join(PASTA, ".git")):
        print("  GitHub: pasta nao e um repositorio git - publicacao ignorada.")
        return
    if not os.path.exists(TOKEN_FILE):
        print("  GitHub: token nao encontrado em %s - publicacao ignorada." % TOKEN_FILE)
        return
    token = open(TOKEN_FILE, encoding="utf-8").read().strip()
    push_url = "https://%s@github.com/%s.git" % (token, REPO_SLUG)
    try:
        _git("add", "-A")
        st = _git("status", "--porcelain").stdout.strip()
        if not st:
            print("  GitHub: nenhuma alteracao para publicar.")
            return
        _git("commit", "-m", msg)
        r = _git("push", push_url, "HEAD:main", check=False)
        if r.returncode == 0:
            print("  GitHub: publicado com sucesso.")
            print("  Online:  https://ricardooliveiralc.github.io/LabecodeOQC/")
        else:
            print("  GitHub: FALHA no push -> %s" % (r.stderr.strip()[:300]))
    except Exception as e:
        print("  GitHub: erro ao publicar -> %s" % e)

# ---------------------------------------------------------------------------
def main():
    if not os.path.exists(TEMPLATE):
        print("ERRO: template.html nao encontrado na pasta.")
        sys.exit(1)
    xlsx = resolver_planilha()
    registros = extrair(xlsx)
    if not registros:
        print("ERRO: nenhuma inspecao encontrada na planilha.")
        sys.exit(1)

    agora = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    tpl = open(TEMPLATE, encoding='utf-8').read()
    html = tpl.replace('__DATA_JSON__', json.dumps(registros, ensure_ascii=False))
    html = html.replace('__GENERATED_AT__', agora)
    open(SAIDA, 'w', encoding='utf-8').write(html)
    shutil.copyfile(SAIDA, SAIDA_WEB)   # index.html para o GitHub Pages

    st = Counter(x['status'] for x in registros)
    print("-" * 52)
    print("Dashboard atualizado com sucesso!")
    print("  Inspecoes:   %d" % len(registros))
    print("  Aprovadas:   %d" % st.get('APROVADO', 0))
    print("  Condicionais:%d" % st.get('CONDICIONAL', 0))
    print("  Reprovadas:  %d" % st.get('REPROVADO', 0))
    print("  Periodo:     %s a %s" % (registros[0]['date'], registros[-1]['date']))
    print("  Gerado em:   %s" % agora)
    print("  Arquivo:     %s" % SAIDA)
    print("-" * 52)
    publicar_github("Atualiza dashboard OQC (auto) - %s" % agora)
    print("-" * 52)

if __name__ == '__main__':
    main()
